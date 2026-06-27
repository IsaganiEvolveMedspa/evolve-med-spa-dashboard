"""
Regenerate data/payroll_schedules.json from data/Supporting Schedules For Railway.xlsx.

Extracts the STATIC payroll inputs the salary model needs (the rest comes from the DB):
  - wages           : Employee Role -> base Hourly Rate            (Hourly Wage sheet)
  - benefits_factor : burden multiplier (taxes/benefits), e.g. 0.125
  - commission_rate : 0.15 (fixed in the salary formula)
  - ffs             : Item Name -> { per_syringe, latest_ffs }     (FFS Schedule sheet)

Keys are lowercased/trimmed for case-insensitive matching against DB values
(job_name for wages, service_name for ffs). Runtime reads only the JSON (stdlib).

Run after refreshing the Excel:
    python scripts/build_payroll_json.py
"""
import json
import os

import openpyxl

HERE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(HERE, "data", "Supporting Schedules For Railway.xlsx")
OUT  = os.path.join(HERE, "data", "payroll_schedules.json")

COMMISSION_RATE = 0.15


def _norm(v) -> str:
    return str(v).strip().lower() if v is not None else ""


def main():
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)

    # ---- Hourly Wage ----
    hw = wb["Hourly Wage"]
    hw_rows = list(hw.iter_rows(values_only=True))
    benefits_factor = float(hw_rows[0][2])          # row0, col2 = 0.125
    wages: dict[str, float] = {}
    for r in hw_rows[2:]:                            # data starts row index 2
        role, rate = r[0], r[1]
        if role is None or rate is None:
            continue
        try:
            wages[_norm(role)] = float(rate)
        except (TypeError, ValueError):
            continue

    # ---- FFS Schedule ----
    ffs_ws = wb["FFS Schedule"]
    ffs: dict[str, dict] = {}
    for i, r in enumerate(ffs_ws.iter_rows(values_only=True)):
        if i == 0:
            continue                                 # header
        item, per_syr, latest = r[2], r[3], r[29]
        if item is None:
            continue
        try:
            latest_val = float(latest) if latest is not None else 0.0
        except (TypeError, ValueError):
            latest_val = 0.0
        ffs[_norm(item)] = {
            "per_syringe": _norm(per_syr) == "y",   # 'Y' -> True, '0' -> False
            "latest_ffs": latest_val,
        }

    out = {
        "benefits_factor": benefits_factor,
        "commission_rate": COMMISSION_RATE,
        "wages": wages,
        "ffs": ffs,
    }
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(out, f, indent=0, sort_keys=True)

    print(f"wrote {len(wages)} roles, {len(ffs)} ffs items, "
          f"benefits_factor={benefits_factor} -> {OUT}")


if __name__ == "__main__":
    main()
