"""
Regenerate data/ad_spend_daily.json from the ad-spend source.

Single source of truth: data/Google Ads.xlsx (re-imported daily). Two sheets feed
MTD Ad Spend and CAC:
  - Google:   'Sheet1' tab
  - Facebook: 'Fb'     tab

Aggregates Google + Facebook ad spend to one chain-level total per calendar day
("Report: Date" -> SUM("Cost: Amount spend")). The API reads the JSON at runtime
(stdlib json only) so production needs no Excel/openpyxl dependency.

Run after refreshing the workbook:
    python scripts/build_ad_spend_json.py
"""
import json
import os
from collections import defaultdict

import openpyxl

HERE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(HERE, "data", "Google Ads.xlsx")
OUT  = os.path.join(HERE, "data", "ad_spend_daily.json")

DATE_COL  = "Report: Date"
SPEND_COL = "Cost: Amount spend"

# Sheet1 = Google export, Fb = Facebook export (both in the daily workbook).
GOOGLE_TAB = "sheet1"
FB_TAB     = "fb"


def _norm_date(d) -> str:
    return d.date().isoformat() if hasattr(d, "date") else str(d)[:10]


def _add_sheet(daily: dict[str, float], wb, tab: str) -> float:
    total = 0.0
    for sheet in wb.sheetnames:
        if sheet.strip().lower() != tab:
            continue
        ws = wb[sheet]
        rows = ws.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            continue
        if DATE_COL not in header or SPEND_COL not in header:
            raise ValueError(f"sheet '{sheet}' missing '{DATE_COL}'/'{SPEND_COL}'")
        di, si = header.index(DATE_COL), header.index(SPEND_COL)
        for r in rows:
            d, spend = r[di], r[si]
            if d is None or spend is None:
                continue
            daily[_norm_date(d)] += float(spend)
            total += float(spend)
    return total


def main():
    daily: dict[str, float] = defaultdict(float)
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    g = _add_sheet(daily, wb, GOOGLE_TAB)
    fb = _add_sheet(daily, wb, FB_TAB)

    daily = {k: round(v, 2) for k, v in sorted(daily.items())}
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(daily, f, indent=0, sort_keys=True)

    print(f"Google (Sheet1): {round(g, 2)}  Facebook (Fb): {round(fb, 2)}")
    print(f"wrote {len(daily)} days, total spend {round(sum(daily.values()), 2)} -> {OUT}")


if __name__ == "__main__":
    main()
